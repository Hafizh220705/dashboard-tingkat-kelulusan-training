"""Utilitas umum untuk konversi CSV dan penggabungan file Excel.

Modul ini tidak bergantung pada Streamlit agar logika pengolahan file dapat
dipakai ulang dan diuji secara terpisah dari antarmuka aplikasi.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import BinaryIO, Iterable
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pandas.errors import MergeError

from config.settings import (
    KOLOM_NILAI_LAPORAN_TRAINING,
    KOLOM_NRP_LAPORAN_TRAINING,
    KOLOM_WAJIB_LAPORAN_TRAINING,
    STATUS_DATA_TIDAK_LENGKAP,
    STATUS_LULUS,
    STATUS_TIDAK_LULUS,
    THRESHOLD_KELULUSAN,
)
from src.data.training_attempts import normalisasi_nilai_pertanyaan


BATAS_BARIS_EXCEL = 1_048_575  # satu baris lain dipakai untuk header
BATAS_KOLOM_EXCEL = 16_384
ENCODING_CSV_DIDUKUNG = ("utf-8-sig", "cp1252", "latin-1")
DELIMITER_CSV_DIDUKUNG = ",;\t|"


class PengolahanFileError(ValueError):
    """Error yang aman ditampilkan kepada pengguna saat pengolahan gagal."""


@dataclass(frozen=True)
class RingkasanKonversiCSV:
    """Ringkasan satu file dalam proses konversi batch."""

    nama_file_asal: str
    nama_file_hasil: str
    jumlah_baris: int
    jumlah_kolom: int


@dataclass(frozen=True)
class RingkasanPencocokanMaster:
    """Statistik hasil pencocokan baris training ke master employee."""

    jumlah_cocok: int
    jumlah_tidak_cocok: int
    jumlah_duplikat_master_dihapus: int


def _nama_file(file_obj: BinaryIO) -> str:
    return str(getattr(file_obj, "name", "file"))


def _baca_seluruh_bytes(file_obj: BinaryIO) -> bytes:
    """Baca file-like object dari awal dan kembalikan posisinya ke awal."""
    try:
        file_obj.seek(0)
        data = file_obj.read()
        file_obj.seek(0)
    except (AttributeError, OSError, ValueError) as exc:
        raise PengolahanFileError("File tidak dapat dibaca.") from exc

    if not isinstance(data, bytes):
        raise PengolahanFileError("Isi file harus berupa data biner.")
    return data


def _validasi_ukuran_excel(df: pd.DataFrame) -> None:
    if len(df) > BATAS_BARIS_EXCEL:
        raise PengolahanFileError(
            f"Data memiliki {len(df):,} baris, melebihi batas Excel "
            f"({BATAS_BARIS_EXCEL:,} baris data)."
        )
    if len(df.columns) > BATAS_KOLOM_EXCEL:
        raise PengolahanFileError(
            f"Data memiliki {len(df.columns):,} kolom, melebihi batas Excel "
            f"({BATAS_KOLOM_EXCEL:,} kolom)."
        )


def _decode_csv(data: bytes, nama_file: str) -> str:
    for encoding in ENCODING_CSV_DIDUKUNG:
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise PengolahanFileError(f"Encoding CSV '{nama_file}' tidak didukung.")


def _deteksi_delimiter(teks: str) -> str:
    sampel = teks[:65_536]
    try:
        return csv.Sniffer().sniff(sampel, delimiters=DELIMITER_CSV_DIDUKUNG).delimiter
    except csv.Error:
        # CSV satu kolom memang tidak mempunyai delimiter yang dapat dideteksi.
        return ","


def baca_csv_umum(uploaded_file: BinaryIO) -> pd.DataFrame:
    """Baca CSV umum tanpa validasi schema dashboard.

    Delimiter koma, titik koma, tab, dan pipe dideteksi otomatis. Encoding yang
    umum ditemui pada file Windows juga ditangani sebagai fallback.
    """
    nama_file = _nama_file(uploaded_file)
    if Path(nama_file).suffix.lower() != ".csv":
        raise PengolahanFileError("Pilih file dengan format .csv.")

    data = _baca_seluruh_bytes(uploaded_file)
    if not data or not data.strip():
        raise PengolahanFileError(f"File CSV '{nama_file}' kosong.")

    teks = _decode_csv(data, nama_file)
    delimiter = _deteksi_delimiter(teks)

    try:
        df = pd.read_csv(StringIO(teks), sep=delimiter)
    except Exception as exc:
        raise PengolahanFileError(
            f"CSV '{nama_file}' tidak dapat dibaca: {exc}"
        ) from exc

    if len(df.columns) == 0:
        raise PengolahanFileError(f"CSV '{nama_file}' tidak memiliki kolom.")

    _validasi_ukuran_excel(df)
    return df


def dataframe_ke_excel(df: pd.DataFrame, nama_sheet: str = "Data") -> bytes:
    """Ubah DataFrame menjadi workbook XLSX siap diunduh."""
    _validasi_ukuran_excel(df)
    output = BytesIO()

    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=nama_sheet)
            worksheet = writer.sheets[nama_sheet]
            worksheet.freeze_panes = "A2"

            if len(df.columns) > 0:
                worksheet.auto_filter.ref = worksheet.dimensions
                warna_header = PatternFill("solid", fgColor="FFD100")
                for cell in worksheet[1]:
                    cell.fill = warna_header
                    cell.font = Font(bold=True, color="1D1D1B")

                # Lebar didasarkan pada header agar hasil langsung mudah dibaca,
                # tanpa memindai seluruh dataset yang mungkin sangat besar.
                for indeks, nama_kolom in enumerate(df.columns, start=1):
                    lebar = min(max(len(str(nama_kolom)) + 2, 10), 40)
                    worksheet.column_dimensions[get_column_letter(indeks)].width = lebar
    except Exception as exc:
        raise PengolahanFileError(f"Gagal membuat file Excel: {exc}") from exc

    return output.getvalue()


def konversi_csv_ke_excel(uploaded_file: BinaryIO) -> tuple[pd.DataFrame, bytes]:
    """Konversi satu CSV menjadi DataFrame dan workbook XLSX."""
    df = baca_csv_umum(uploaded_file)
    return df, dataframe_ke_excel(df)


def _nama_hasil_konversi_unik(nama_file: str, nama_terpakai: set[str]) -> str:
    stem = Path(nama_file).stem or "hasil"
    kandidat = f"{stem}.xlsx"
    urutan = 2

    while kandidat.casefold() in nama_terpakai:
        kandidat = f"{stem}_{urutan}.xlsx"
        urutan += 1

    nama_terpakai.add(kandidat.casefold())
    return kandidat


def konversi_banyak_csv_ke_zip(
    uploaded_files: Iterable[BinaryIO],
) -> tuple[list[RingkasanKonversiCSV], bytes]:
    """Konversi beberapa CSV dan kemas seluruh XLSX dalam satu file ZIP."""
    files = list(uploaded_files)
    if not files:
        raise PengolahanFileError("Upload minimal satu file CSV untuk dikonversi.")

    output = BytesIO()
    ringkasan: list[RingkasanKonversiCSV] = []
    nama_terpakai: set[str] = set()

    with ZipFile(output, mode="w", compression=ZIP_DEFLATED) as arsip:
        for uploaded_file in files:
            df, hasil_excel = konversi_csv_ke_excel(uploaded_file)
            nama_asal = _nama_file(uploaded_file)
            nama_hasil = _nama_hasil_konversi_unik(nama_asal, nama_terpakai)
            arsip.writestr(nama_hasil, hasil_excel)
            ringkasan.append(
                RingkasanKonversiCSV(
                    nama_file_asal=nama_asal,
                    nama_file_hasil=nama_hasil,
                    jumlah_baris=len(df),
                    jumlah_kolom=len(df.columns),
                )
            )

    return ringkasan, output.getvalue()


def _bersihkan_nama_kolom_excel(df: pd.DataFrame, nama_file: str) -> pd.DataFrame:
    hasil = df.copy()
    hasil.columns = [" ".join(str(kolom).split()) for kolom in hasil.columns]

    kolom_duplikat = hasil.columns[hasil.columns.duplicated()].unique().tolist()
    if kolom_duplikat:
        raise PengolahanFileError(
            f"File '{nama_file}' memiliki nama kolom duplikat: "
            f"{', '.join(map(str, kolom_duplikat))}."
        )
    return hasil


def _nama_kolom_sumber(kolom_terpakai: Iterable[object]) -> str:
    nama_terpakai = {str(kolom) for kolom in kolom_terpakai}
    kandidat = "SUMBER_FILE"
    urutan = 2
    while kandidat in nama_terpakai:
        kandidat = f"SUMBER_FILE_{urutan}"
        urutan += 1
    return kandidat


def _normalisasi_kunci_nrp(series: pd.Series) -> pd.Series:
    """Samakan representasi NRP dari teks dan angka untuk kebutuhan join."""
    hasil = (
        series.astype("string")
        .str.strip()
        .str.replace(r"\.0+$", "", regex=True)
        .str.replace(r"\s+", "", regex=True)
        .str.upper()
    )
    hasil = hasil.mask(hasil.isin(["", "NAN", "NONE", "<NA>"]))

    # Excel kadang menghilangkan nol di depan saat membaca NRP sebagai angka.
    # Untuk NRP yang seluruhnya numerik, 00123 dan 123 diperlakukan sama.
    mask_numerik = hasil.str.fullmatch(r"\d+", na=False)
    hasil.loc[mask_numerik] = hasil.loc[mask_numerik].str.lstrip("0").replace("", "0")
    return hasil


def _nilai_grade_numerik(series: pd.Series) -> pd.Series:
    """Parse grade skala 100, termasuk angka dengan koma desimal atau persen."""
    teks = (
        series.astype("string")
        .str.strip()
        .str.replace("%", "", regex=False)
        .str.replace(",", ".", regex=False)
    )
    return pd.to_numeric(teks, errors="coerce")


def siapkan_laporan_training(
    df: pd.DataFrame,
    threshold: float = THRESHOLD_KELULUSAN,
) -> pd.DataFrame:
    """Bentuk NRP dan status kelulusan dari template laporan training."""
    kolom_hilang = [
        kolom for kolom in KOLOM_WAJIB_LAPORAN_TRAINING if kolom not in df.columns
    ]
    if kolom_hilang:
        raise PengolahanFileError(
            "Template laporan training tidak lengkap. Kolom yang hilang: "
            f"{', '.join(kolom_hilang)}."
        )

    hasil = normalisasi_nilai_pertanyaan(df)
    hasil["NRP_ID"] = _normalisasi_kunci_nrp(
        hasil[KOLOM_NRP_LAPORAN_TRAINING]
    )
    hasil["NILAI_FINAL"] = _nilai_grade_numerik(
        hasil[KOLOM_NILAI_LAPORAN_TRAINING]
    )

    hasil["RESULT_FINAL"] = STATUS_DATA_TIDAK_LENGKAP
    mask_nilai_ada = hasil["NILAI_FINAL"].notna()
    hasil.loc[
        mask_nilai_ada & (hasil["NILAI_FINAL"] >= threshold),
        "RESULT_FINAL",
    ] = STATUS_LULUS
    hasil.loc[
        mask_nilai_ada & (hasil["NILAI_FINAL"] < threshold),
        "RESULT_FINAL",
    ] = STATUS_TIDAK_LULUS
    return hasil


def hubungkan_laporan_dengan_master(
    df_training: pd.DataFrame,
    df_master: pd.DataFrame,
    kolom_nrp_master: str = "NRP",
) -> tuple[pd.DataFrame, RingkasanPencocokanMaster]:
    """Tambahkan atribut master ke training melalui NRP tanpa melipatgandakan baris."""
    if "NRP_ID" not in df_training.columns:
        raise PengolahanFileError(
            "Data training belum memiliki NRP_ID. Siapkan laporan training terlebih dahulu."
        )
    if kolom_nrp_master not in df_master.columns:
        raise PengolahanFileError(
            f"Master employee tidak memiliki kolom '{kolom_nrp_master}'."
        )

    training = df_training.copy()
    master = df_master.copy()
    training["NRP_ID"] = _normalisasi_kunci_nrp(training["NRP_ID"])
    nama_kunci_master = "__NRP_MASTER_KEY__"
    while nama_kunci_master in master.columns or nama_kunci_master in training.columns:
        nama_kunci_master = f"_{nama_kunci_master}"

    master[nama_kunci_master] = _normalisasi_kunci_nrp(master[kolom_nrp_master])
    master = master[master[nama_kunci_master].notna()].copy()
    jumlah_sebelum_dedup = len(master)
    master = master.drop_duplicates(subset=[nama_kunci_master], keep="first")
    jumlah_duplikat = jumlah_sebelum_dedup - len(master)

    nama_indikator = "__STATUS_PENCOCOKAN_MASTER__"
    while nama_indikator in master.columns or nama_indikator in training.columns:
        nama_indikator = f"_{nama_indikator}"

    try:
        hasil = training.merge(
            master,
            how="left",
            left_on="NRP_ID",
            right_on=nama_kunci_master,
            suffixes=("", "_MASTER"),
            indicator=nama_indikator,
            validate="many_to_one",
        )
    except (KeyError, MergeError, ValueError) as exc:
        raise PengolahanFileError(
            f"Gagal menghubungkan laporan training dengan master: {exc}"
        ) from exc

    mask_cocok = hasil[nama_indikator] == "both"
    ringkasan = RingkasanPencocokanMaster(
        jumlah_cocok=int(mask_cocok.sum()),
        jumlah_tidak_cocok=int((~mask_cocok).sum()),
        jumlah_duplikat_master_dihapus=jumlah_duplikat,
    )
    hasil = hasil.drop(columns=[nama_kunci_master, nama_indikator])
    return hasil, ringkasan


def gabungkan_file_excel(
    uploaded_files: Iterable[BinaryIO],
    tambah_sumber_file: bool = False,
    kolom_wajib: Iterable[str] | None = None,
) -> pd.DataFrame:
    """Gabungkan sheet pertama dari beberapa workbook secara vertikal.

    Urutan kolom mengikuti file pertama, lalu kolom baru dari file berikutnya.
    Jika suatu file tidak mempunyai sebuah kolom, nilainya dibiarkan kosong.
    """
    files = list(uploaded_files)
    if len(files) < 2:
        raise PengolahanFileError("Upload minimal dua file Excel untuk digabungkan.")

    dataframes: list[pd.DataFrame] = []
    nama_files: list[str] = []
    daftar_kolom_wajib = list(kolom_wajib or [])

    for uploaded_file in files:
        nama_file = _nama_file(uploaded_file)
        if Path(nama_file).suffix.lower() != ".xlsx":
            raise PengolahanFileError(
                f"File '{nama_file}' bukan .xlsx dan tidak dapat digabungkan."
            )

        data = _baca_seluruh_bytes(uploaded_file)
        if not data:
            raise PengolahanFileError(f"File Excel '{nama_file}' kosong.")

        try:
            df = pd.read_excel(BytesIO(data), sheet_name=0, engine="openpyxl")
        except Exception as exc:
            raise PengolahanFileError(
                f"Excel '{nama_file}' tidak dapat dibaca: {exc}"
            ) from exc

        if len(df.columns) == 0:
            raise PengolahanFileError(
                f"Sheet pertama pada '{nama_file}' tidak memiliki kolom."
            )

        df = _bersihkan_nama_kolom_excel(df, nama_file)
        kolom_hilang = [kolom for kolom in daftar_kolom_wajib if kolom not in df.columns]
        if kolom_hilang:
            raise PengolahanFileError(
                f"File '{nama_file}' tidak sesuai template. Kolom yang hilang: "
                f"{', '.join(kolom_hilang)}."
            )

        dataframes.append(df)
        nama_files.append(nama_file)

    if tambah_sumber_file:
        semua_kolom = [kolom for df in dataframes for kolom in df.columns]
        kolom_sumber = _nama_kolom_sumber(semua_kolom)
        for df, nama_file in zip(dataframes, nama_files):
            df[kolom_sumber] = nama_file

    hasil = pd.concat(dataframes, ignore_index=True, sort=False)
    if len(hasil) == 0:
        raise PengolahanFileError(
            "Semua sheet pertama yang dipilih kosong dan tidak memiliki baris data."
        )

    _validasi_ukuran_excel(hasil)
    return hasil


def gabungkan_excel_ke_bytes(
    uploaded_files: Iterable[BinaryIO],
    tambah_sumber_file: bool = False,
) -> tuple[pd.DataFrame, bytes]:
    """Gabungkan workbook lalu hasilkan DataFrame dan file XLSX."""
    hasil = gabungkan_file_excel(uploaded_files, tambah_sumber_file)
    return hasil, dataframe_ke_excel(hasil, nama_sheet="Data Gabungan")
